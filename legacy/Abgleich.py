from openpyxl import load_workbook, Workbook
from pathlib import Path, PureWindowsPath
from time import sleep

# Konstanten
NETZWERKPFAD = str(PureWindowsPath(r"Z:\KKK\Fachbereiche\TKW\Allgemein\ReVS Arbeitsordner\12_ReVS-AVK-Abgleich"))
REVS_DATEI = "export.xlsx"
AVK_DATEINAMEN = ["AVK.xlsx", "Avk.xlsx", "avk.xlsx"]
ABGLEICH_DATEI = "Abgleich.xlsx"

SPALTEN_REVS = ["Verpackungs-ID", "Standort", "Reststoff-ID", "Nettomasse/kg"]
SPALTEN_AVK = ["Behälternummer", "Lagerort/Absender", "Abfallmasse [kg]", "Individuelle ID"]
FEHLER_HEADER = [
    "Verpackungs-ID", "Reststoff-ID", "Individuelle ID",
    "Standort ReVS", "Standort AVK",
    "ReVS Nettomasse/kg", "AVK Abfallmasse [kg]",
]

ORT_KKK = "KKK"
STANDORT_AN_AVK = "An AVK übergeben"
STANDORT_ZW6 = "ZW6-1"
STANDORT_CONTAINER = "Containerstellplatz-ÜB"
CONTAINER_BEZEICHNUNG = 'CONTAINER 20"'
GEBINDETYP_DREISTELLIG = ("V", "E")
GEBINDETYP_DREISTELLIG_ZWEI = ("MH", "SO", "AK", "KE")


def lade_dateien(pfad):
    """Lädt ReVS- und AVK-Workbooks aus dem angegebenen Pfad."""
    revs_pfad = Path(pfad, REVS_DATEI)
    if not revs_pfad.exists():
        raise FileNotFoundError(f"ReVS-Datei nicht gefunden: {revs_pfad}")
    revs = load_workbook(str(revs_pfad)).active

    avk_sheet = None
    for dateiname in AVK_DATEINAMEN:
        avk_pfad = Path(pfad, dateiname)
        if avk_pfad.exists():
            avk_sheet = load_workbook(str(avk_pfad)).active
            break
    if avk_sheet is None:
        raise FileNotFoundError(f"AVK-Datei nicht gefunden in: {pfad}")

    return revs, avk_sheet


def lese_spalten_indizes(sheet, spalten_namen):
    """Ermittelt die Spaltenindizes für die angegebenen Spaltennamen aus dem Header."""
    header = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    indizes = {}
    for name in spalten_namen:
        if name not in header:
            raise ValueError(
                f"Spalte '{name}' nicht im Header gefunden. Verfügbare Spalten: {header}"
            )
        indizes[name] = header.index(name) + 1
    return indizes


def _berechne_avk_standort(standort):
    """Berechnet den AVK-Übersetzungsstandort aus dem ReVS-Standort."""
    if not standort:
        return ""
    if standort[0:2] == "EX":
        return ORT_KKK
    if standort == STANDORT_ZW6:
        return "W 06"
    if standort == STANDORT_CONTAINER:
        return CONTAINER_BEZEICHNUNG
    if standort[0:2] in ("A-", "AL", "ST") or standort[0:1] == "E":
        z_pos = standort.find("Z")
        if z_pos == -1:
            return standort
        return standort[z_pos + 1: z_pos + 2] + " " + standort[z_pos + 2: z_pos + 7]
    return ""


def vergleiche_gebinde(revs, avk, idx_revs, idx_avk):
    """Führt den Abgleich der Gebinde zwischen ReVS und AVK durch."""
    aufgenommen_avk = set()
    aufgenommen_revs = set()
    ort_kkk_vorhanden = False

    # AVK-Gebinde einlesen
    gebinde_avk = []
    for zeile in range(1, avk.max_row + 1):
        zellwert = avk.cell(zeile, idx_avk["Behälternummer"]).value
        if zellwert is None or zellwert == "":
            continue
        zellwert_str = str(zellwert)

        leerzeichen_pos = zellwert_str.find(" ")
        gebinde_typ = zellwert_str[0:leerzeichen_pos] if leerzeichen_pos != -1 else zellwert_str

        stern_pos = zellwert_str.find("*")
        if stern_pos != -1:
            gebinde_nummer = zellwert_str[0:stern_pos]
        elif zellwert_str[0:3] == ORT_KKK:
            gebinde_typ = ORT_KKK
            gebinde_nummer = zellwert_str
        else:
            gebinde_nummer = zellwert_str

        lagerort = avk.cell(zeile, idx_avk["Lagerort/Absender"]).value
        lagerort = lagerort if lagerort is not None else ""
        masse = avk.cell(zeile, idx_avk["Abfallmasse [kg]"]).value
        masse = masse if masse is not None else ""
        ind_id = avk.cell(zeile, idx_avk["Individuelle ID"]).value
        ind_id = ind_id if ind_id is not None else ""

        if gebinde_typ:
            aufgenommen_avk.add(gebinde_typ)
        if str(lagerort) == ORT_KKK:
            ort_kkk_vorhanden = True

        gebinde_avk.append({
            "typ": gebinde_typ,
            "nummer": gebinde_nummer,
            "lagerort": lagerort,
            "masse": masse,
            "ind_id": ind_id,
            "vorhanden": False,
        })

    # ReVS-Gebinde einlesen
    gebinde_revs = []
    for zeile in range(1, revs.max_row + 1):
        verpackungs_id = revs.cell(zeile, idx_revs["Verpackungs-ID"]).value
        if verpackungs_id is None or verpackungs_id == "":
            continue
        verpackungs_id_str = str(verpackungs_id)

        minus_pos = verpackungs_id_str.find("-")
        gebinde_typ = verpackungs_id_str[0:minus_pos] if minus_pos != -1 else verpackungs_id_str

        if gebinde_typ not in aufgenommen_avk:
            continue

        # Gebindenummer bestimmen
        if minus_pos != -1:
            if gebinde_typ[0:1] in GEBINDETYP_DREISTELLIG or gebinde_typ[0:2] in GEBINDETYP_DREISTELLIG_ZWEI:
                gebinde_nummer = gebinde_typ + " " + verpackungs_id_str[minus_pos + 2: minus_pos + 5]
            elif gebinde_typ[0:3] == ORT_KKK:
                gebinde_nummer = ""
            else:
                gebinde_nummer = gebinde_typ + " " + verpackungs_id_str[minus_pos + 1: minus_pos + 7]
        else:
            gebinde_nummer = gebinde_typ

        standort = revs.cell(zeile, idx_revs["Standort"]).value
        standort = str(standort) if standort is not None else ""
        reststoff_id = revs.cell(zeile, idx_revs["Reststoff-ID"]).value
        reststoff_id = reststoff_id if reststoff_id is not None else ""
        nettomasse = revs.cell(zeile, idx_revs["Nettomasse/kg"]).value
        nettomasse = nettomasse if nettomasse is not None else ""

        if gebinde_typ:
            aufgenommen_revs.add(gebinde_typ)

        avk_standort = _berechne_avk_standort(standort)

        gebinde_revs.append({
            "typ": gebinde_typ,
            "nummer": gebinde_nummer,
            "standort": standort,
            "reststoff_id": reststoff_id,
            "nettomasse": nettomasse,
            "avk_standort": avk_standort,
            "vorhanden": False,
        })

    # Abgleich durchführen
    fehler = []
    for revs_eintrag in gebinde_revs:
        for avk_eintrag in gebinde_avk:
            if avk_eintrag["vorhanden"]:
                continue
            if revs_eintrag["nummer"] == avk_eintrag["nummer"]:
                revs_eintrag["vorhanden"] = True
                avk_eintrag["vorhanden"] = True

                abweichungen = 0
                ind_id_out = avk_eintrag["ind_id"]
                standort_avk_out = avk_eintrag["lagerort"]
                masse_avk_out = avk_eintrag["masse"]

                if str(revs_eintrag["reststoff_id"]) == str(avk_eintrag["ind_id"]):
                    ind_id_out = "stimmt mit ReVS"
                else:
                    abweichungen += 1

                if revs_eintrag["avk_standort"] == str(avk_eintrag["lagerort"]):
                    standort_avk_out = "stimmt mit ReVS"
                else:
                    abweichungen += 1

                if str(revs_eintrag["nettomasse"]) == str(avk_eintrag["masse"]):
                    masse_avk_out = "stimmt mit ReVS"
                else:
                    abweichungen += 1

                if abweichungen > 0:
                    fehler.append([
                        revs_eintrag["nummer"],
                        revs_eintrag["reststoff_id"],
                        ind_id_out,
                        revs_eintrag["standort"],
                        standort_avk_out,
                        revs_eintrag["nettomasse"],
                        masse_avk_out,
                    ])
                break

        # ReVS-Gebinde nicht im AVK gefunden
        # KKK-Gebinde nur melden, wenn KKK auch im AVK-Auszug enthalten ist
        if (
            not revs_eintrag["vorhanden"]
            and revs_eintrag["typ"] in aufgenommen_avk
            and not (revs_eintrag["avk_standort"] == ORT_KKK and not ort_kkk_vorhanden)
            and revs_eintrag["standort"] != STANDORT_AN_AVK
        ):
            fehler.append([
                revs_eintrag["nummer"],
                revs_eintrag["reststoff_id"],
                "",
                revs_eintrag["standort"],
                "nicht im AVK gelistet",
                revs_eintrag["nettomasse"],
                "",
            ])

    # AVK-Gebinde nicht im ReVS gefunden
    for avk_eintrag in gebinde_avk:
        if not avk_eintrag["vorhanden"] and avk_eintrag["typ"] in aufgenommen_revs:
            fehler.append([
                avk_eintrag["nummer"],
                "nicht im ReVS gelistet",
                avk_eintrag["ind_id"],
                "",
                avk_eintrag["lagerort"],
                "",
                avk_eintrag["masse"],
            ])

    return fehler


def schreibe_ergebnis(pfad, fehler):
    """Schreibt die Fehlerliste in eine neue Excel-Datei."""
    ergebnis = Workbook()
    ws = ergebnis.active
    ws.cell(1, 1).value = "Fehlerhafte und fehlende Gebinde"
    ws.append(FEHLER_HEADER)
    for zeile in fehler:
        ws.append(zeile)
    ausgabe_pfad = str(Path(pfad, ABGLEICH_DATEI))
    ergebnis.save(ausgabe_pfad)
    ergebnis.close()


def main():
    print("erstellt von Gregor Schuboth")
    sleep(1)
    pfad = NETZWERKPFAD

    try:
        revs, avk = lade_dateien(pfad)
    except FileNotFoundError as e:
        print(f"Fehler beim Laden der Dateien: {e}")
        return

    try:
        idx_revs = lese_spalten_indizes(revs, SPALTEN_REVS)
        idx_avk = lese_spalten_indizes(avk, SPALTEN_AVK)
    except ValueError as e:
        print(f"Fehler beim Einlesen der Spalten: {e}")
        return

    fehler = vergleiche_gebinde(revs, avk, idx_revs, idx_avk)
    schreibe_ergebnis(pfad, fehler)
    print(f"Abgleich abgeschlossen. {len(fehler)} Abweichungen gefunden.")


if __name__ == "__main__":
    main()
