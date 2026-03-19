"""
AVK-vs-ReVS Datei-Handler
Lädt und schreibt Excel-Dateien für den Abgleich.
"""
from pathlib import Path
from typing import Any, Iterator, List, Tuple

from openpyxl import load_workbook, Workbook

from abgleich_logic import FEHLER_HEADER


class _SheetData:
    """Wrapper für read_only geladene Worksheet-Daten (kompatibel mit openpyxl-Sheet-Interface)."""

    class _Cell:
        __slots__ = ("value",)

        def __init__(self, value: Any) -> None:
            self.value = value

    def __init__(self, rows: List[tuple]) -> None:
        self._rows = rows

    @property
    def max_row(self) -> int:
        return len(self._rows)

    @property
    def max_column(self) -> int:
        return max((len(r) for r in self._rows), default=0)

    def cell(self, row: int, col: int) -> "_SheetData._Cell":
        try:
            return self._Cell(self._rows[row - 1][col - 1])
        except IndexError:
            return self._Cell(None)

    def iter_rows(self, min_row: int = 1, max_row: int = None, values_only: bool = False) -> Iterator:
        end = max_row if max_row is not None else len(self._rows)
        for row_idx in range(min_row - 1, end):
            if row_idx >= len(self._rows):
                break
            yield self._rows[row_idx]


def _lade_sheet_daten(pfad: Path) -> _SheetData:
    """Öffnet ein Workbook im read_only-Modus, liest alle Daten ein und schließt es."""
    wb = load_workbook(str(pfad), read_only=True, data_only=True)
    try:
        rows = [tuple(row) for row in wb.active.iter_rows(values_only=True)]
    finally:
        wb.close()
    return _SheetData(rows)


def lade_dateien(
    pfad: str,
    revs_datei: str,
    avk_dateinamen: List[str],
) -> Tuple[_SheetData, _SheetData]:
    """Lädt ReVS- und AVK-Daten aus dem angegebenen Pfad."""
    revs_pfad = Path(pfad, revs_datei)
    if not revs_pfad.exists():
        raise FileNotFoundError(f"ReVS-Datei nicht gefunden: {revs_pfad}")
    revs = _lade_sheet_daten(revs_pfad)

    avk_sheet = None
    for dateiname in avk_dateinamen:
        avk_pfad = Path(pfad, dateiname)
        if avk_pfad.exists():
            avk_sheet = _lade_sheet_daten(avk_pfad)
            break
    if avk_sheet is None:
        raise FileNotFoundError(f"AVK-Datei nicht gefunden in: {pfad}")

    return revs, avk_sheet


def schreibe_ergebnis(
    pfad: str,
    abgleich_datei: str,
    fehler: List[List[Any]],
) -> str:
    """Schreibt die Fehlerliste in eine neue Excel-Datei und gibt den Ausgabepfad zurück."""
    ergebnis = Workbook()
    ws = ergebnis.active
    ws.cell(1, 1).value = "Fehlerhafte und fehlende Gebinde"
    ws.append(FEHLER_HEADER)
    for zeile in fehler:
        ws.append(zeile)
    ausgabe_pfad = str(Path(pfad, abgleich_datei))
    ergebnis.save(ausgabe_pfad)
    ergebnis.close()
    return ausgabe_pfad
